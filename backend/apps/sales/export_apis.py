"""
Excel export for sales.

Output mirrors the shop's own «savdo » / «nomerla » / «davomat » workbook
byte-for-byte so a manager can download and merge it back into their
existing spreadsheet workflow:

  - sheet names carry the trailing space the source uses ("savdo ", not "savdo")
  - row 0 mixes the FIRST day's date marker (col A) with the trailing-space
    header words ("ime ", "mijoz ", "Hamkorlar ", "ishchila ", "Daplata ",
    "Rasxodlar ", "izoh") in the same row layout as the source
  - split payments serialize back as "Birzum+Hamroh" / "5300000+6900000"
    on a single row, restoring the shape of the original input
  - phone numbers, IMEI short codes, amounts written as numbers (not
    strings) when single, strings only when concatenated with "+"
  - empty `davomat ` sheet is added because the source template has one

Filters (`search`, `operator`, `channel`, `date_from`, `date_to`) reuse
`sale_list` so the workbook contains exactly the rows the manager sees
in the UI list at the moment of click.
"""

from __future__ import annotations

from collections import defaultdict
from collections.abc import Iterable

from django.utils import timezone
from django.utils.dateparse import parse_datetime
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView

from apps.common.excel import (
    HEADER_FILL,
    HEADER_FONT,
    MONEY_FMT,
    new_workbook,
    workbook_response,
)
from apps.operators.models import Operator, OperatorStatus
from apps.users.permissions import IsTeamLeadOrManagerReadOnly

from .models import Sale
from .selectors import sale_list

# Trailing spaces are intentional — that's how the source workbook stores
# them and the importer's sheet-lookup is case- and whitespace-insensitive.
SAVDO_SHEET_NAME = "savdo "
NOMERLA_SHEET_NAME = "nomerla "
DAVOMAT_SHEET_NAME = "davomat "

# Header words for row 0 columns B..Q. Column A on row 0 holds the first
# date marker (so the very first day's sales fit directly under it without
# a separate marker row). Trailing spaces preserved to match the source.
SAVDO_HEADER_BY_COL: dict[int, str | None] = {
    1: "ime ",
    2: "mijoz ",
    3: "mijoz nomer",
    4: "Hamkorlar ",
    5: None,            # F — amount column, source merged the header
    6: "ishchila ",
    7: None,
    8: "Daplata ",
    15: "Rasxodlar ",
    16: "izoh",
}

SAVDO_COL_COUNT = 17  # A..Q
SAVDO_DATE_FMT = "MM.DD.YY"
DATE_MARKER_FILL = PatternFill("solid", fgColor="FEF3C7")
DATE_MARKER_FONT = Font(bold=True)


def _split_comment(comment: str) -> dict:
    """
    Pull the prefixed parts back out of comments the importer/UI stored
    in the shape «клиент: X | тел: Y | предоплата: Z | расход: W | rest».
    Anything that doesn't match a prefix is concatenated into `rest`.
    """
    out = {"client": "", "phone": "", "downpayment": "", "expense": "", "rest": []}
    for part in (comment or "").split(" | "):
        s = part.strip()
        if not s:
            continue
        for key, prefix in (
            ("client", "клиент: "),
            ("phone", "тел: "),
            ("downpayment", "предоплата: "),
            ("expense", "расход: "),
        ):
            if s.startswith(prefix):
                out[key] = s[len(prefix):].strip()
                break
        else:
            out["rest"].append(s)
    out["rest"] = " | ".join(out["rest"])
    return out


def _maybe_number(val: str):
    """
    Return int / float if the string is a plain number, else the trimmed
    string. The source workbook stores phones and amounts as numbers (not
    strings), so we restore them on export when round-tripping back.
    """
    s = (val or "").strip()
    if not s:
        return None
    if s.isdigit():
        try:
            return int(s)
        except ValueError:
            return s
    # plain float like "50000.0" — kept as int if there's no fractional part,
    # else float so xlsx-stored numeric formatting still works.
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except (ValueError, TypeError):
        return s


def _joined_partner_names(sale: Sale) -> str:
    lines = list(sale.partner_lines.select_related("partner").all())
    if lines:
        return "+".join(line.partner.name for line in lines)
    return sale.channel.name if sale.channel_id else ""


def _joined_partner_amounts(sale: Sale):
    lines = list(sale.partner_lines.all())
    if len(lines) > 1:
        return "+".join(str(int(line.amount)) for line in lines)
    if lines:
        return float(lines[0].amount)
    return float(sale.amount)


def _joined_operator_names(sale: Sale) -> str:
    lines = list(sale.operator_lines.select_related("operator").all())
    if lines:
        return "+".join(line.operator.full_name for line in lines)
    return sale.operator.full_name if sale.operator_id else ""


def _sale_to_savdo_row(sale: Sale) -> list:
    parsed = _split_comment(sale.comment or "")
    row: list = [None] * SAVDO_COL_COUNT
    row[0] = sale.phone_model or ""
    # Short codes stay short, real 15-digit IMEIs stay 15-digit. Match
    # the source convention of numbers-as-numbers when possible.
    row[1] = _maybe_number(sale.imei)
    row[2] = parsed["client"] or None
    row[3] = _maybe_number(parsed["phone"])
    row[4] = _joined_partner_names(sale)
    row[5] = _joined_partner_amounts(sale)
    row[6] = _joined_operator_names(sale)
    row[7] = _maybe_number(parsed["downpayment"])
    row[15] = _maybe_number(parsed["expense"])
    row[16] = parsed["rest"] or None
    return row


def _style_header_cell(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_date_marker_cell(cell):
    cell.fill = DATE_MARKER_FILL
    cell.font = DATE_MARKER_FONT
    cell.number_format = SAVDO_DATE_FMT


def _write_savdo_sheet(wb, sales: Iterable[Sale]) -> None:
    ws = wb.create_sheet(SAVDO_SHEET_NAME)

    tz = timezone.get_current_timezone()
    by_date: dict = defaultdict(list)
    for s in sales:
        by_date[s.sold_at.astimezone(tz).date()].append(s)
    days = sorted(by_date.keys())

    # ---- Row 0: first date marker in col A + header words in B..Q ----
    header_row = [None] * SAVDO_COL_COUNT
    if days:
        header_row[0] = days[0]
    for col_idx, word in SAVDO_HEADER_BY_COL.items():
        header_row[col_idx] = word
    ws.append(header_row)
    if days:
        _style_date_marker_cell(ws.cell(row=1, column=1))
    for col_idx, word in SAVDO_HEADER_BY_COL.items():
        if word:
            _style_header_cell(ws.cell(row=1, column=col_idx + 1))

    # ---- First day's sales fit directly under row 0 (no extra marker) ----
    if days:
        for sale in by_date[days[0]]:
            ws.append(_sale_to_savdo_row(sale))

    # ---- Remaining days: standalone date marker rows + sale rows ----
    for d in days[1:]:
        marker_row = [None] * SAVDO_COL_COUNT
        marker_row[0] = d
        ws.append(marker_row)
        _style_date_marker_cell(ws.cell(row=ws.max_row, column=1))

        for sale in by_date[d]:
            ws.append(_sale_to_savdo_row(sale))

    # ---- Column widths (don't write a "TOTAL" row — source doesn't have one) ----
    for col_idx in range(1, SAVDO_COL_COUNT + 1):
        letter = get_column_letter(col_idx)
        max_len = len(SAVDO_HEADER_BY_COL.get(col_idx - 1, "") or "")
        for cell in ws[letter]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 50)

    # Money format on the amount column (only on numeric cells — strings stay).
    for col in ws.iter_cols(min_col=6, max_col=6, min_row=2):
        for cell in col:
            if isinstance(cell.value, int | float):
                cell.number_format = MONEY_FMT

    ws.freeze_panes = "A2"


def _write_nomerla_sheet(wb) -> None:
    ws = wb.create_sheet(NOMERLA_SHEET_NAME)
    ws.append(["Isim ", "shaxsiy ", "kampaniya"])
    for cell in ws[1]:
        _style_header_cell(cell)
    for op in (
        Operator.objects.exclude(status=OperatorStatus.INACTIVE).order_by("full_name")
    ):
        phone = _maybe_number(op.phone or "")
        ws.append([op.full_name, phone, None])
    for col in (1, 2, 3):
        ws.column_dimensions[get_column_letter(col)].width = 24
    ws.freeze_panes = "A2"


def _write_davomat_sheet(wb) -> None:
    """Empty attendance sheet — kept so the workbook structure matches 1:1."""
    wb.create_sheet(DAVOMAT_SHEET_NAME)


class SaleExportApi(APIView):
    permission_classes = [IsTeamLeadOrManagerReadOnly]

    def get(self, request):
        params = request.query_params
        qs = (
            sale_list(
                search=params.get("search"),
                operator_id=params.get("operator") or None,
                channel_id=params.get("channel") or None,
                date_from=parse_datetime(params.get("date_from"))
                if params.get("date_from")
                else None,
                date_to=parse_datetime(params.get("date_to"))
                if params.get("date_to")
                else None,
            )
            .prefetch_related("operator_lines__operator", "partner_lines__partner")
            .order_by("sold_at", "id")
        )

        wb = new_workbook()
        _write_savdo_sheet(wb, list(qs))
        _write_nomerla_sheet(wb)
        _write_davomat_sheet(wb)
        return workbook_response(wb, "naffcrm-savdo.xlsx")

"""
Excel-экспорт: итоговая строка на «savdo » + лист «svodka»
(кто сколько продал / что продавали / общая сумма).
"""

from __future__ import annotations

from decimal import Decimal

import pytest
from django.contrib.auth import get_user_model
from django.utils import timezone
from openpyxl import load_workbook
from io import BytesIO
from rest_framework.test import APIClient

from apps.catalog.models import Channel
from apps.operators.models import Operator, OperatorStatus
from apps.sales.models import Sale, SaleOperator, SaleStatus
from apps.users.models import Profile, Role

User = get_user_model()


@pytest.fixture
def mgr_client(db):
    u = User.objects.create_user(username="mgr", password="x")
    Profile.objects.create(user=u, role=Role.MANAGER)
    c = APIClient()
    c.force_authenticate(u)
    return c


@pytest.fixture
def dataset(db):
    ch = Channel.objects.create(name="Наличные")
    a = Operator.objects.create(full_name="Alice", status=OperatorStatus.ACTIVE)
    b = Operator.objects.create(full_name="Bob", status=OperatorStatus.ACTIVE)
    for op, amount, model in ((a, 1000, "iPhone 15"), (a, 2000, "Samsung A05"), (b, 3000, "iPhone 15")):
        s = Sale.objects.create(
            imei=f"35{amount}",
            phone_model=model,
            amount=Decimal(amount),
            operator=op,
            channel=ch,
            status=SaleStatus.CONFIRMED,
            sold_at=timezone.now(),
        )
        SaleOperator.objects.create(sale=s, operator=op, amount=Decimal(amount))
    return a, b


@pytest.mark.django_db
def test_export_has_total_row_and_svodka(mgr_client, dataset):
    r = mgr_client.get("/api/sales/export.xlsx")
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    assert "svodka" in wb.sheetnames

    # --- итоговая строка на savdo ---
    ws = wb["savdo "]
    rows = list(ws.iter_rows(values_only=True))
    total_row = next(r for r in rows if r and r[0] == "ИТОГО:")
    assert total_row[1] == "3 ta savdo"
    assert total_row[18] == 6000.0  # net = 6000 - 0 скидок
    # Колонка G (оператор, idx 6) пустая — иначе re-import съел бы строку.
    assert total_row[6] is None

    # --- svodka ---
    sv = wb["svodka"]
    flat = [tuple(r) for r in sv.iter_rows(values_only=True)]
    assert ("Всего продаж", 3) == flat[1][:2]
    assert ("Общая сумма", 6000.0) == flat[2][:2]
    ops = {r[0]: (r[1], r[2]) for r in flat if r[0] in ("Alice", "Bob")}
    assert ops["Alice"] == (2, 3000.0)
    assert ops["Bob"] == (1, 3000.0)
    models = {r[0]: (r[1], r[2]) for r in flat if r[0] in ("iPhone 15", "Samsung A05")}
    assert models["iPhone 15"] == (2, 4000.0)


@pytest.mark.django_db
def test_export_empty_has_no_total_row(mgr_client):
    r = mgr_client.get("/api/sales/export.xlsx")
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    ws = wb["savdo "]
    rows = list(ws.iter_rows(values_only=True))
    assert not any(r and r[0] == "ИТОГО:" for r in rows)

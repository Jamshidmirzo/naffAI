"""
Excel importer for the «savdo» (sales) ledger our shop uses.

Source format (sheet `savdo`):

  Col A  — first cell is either a `datetime` section marker for the day's
           sales, or the phone model name for a sale row.
  Col B  — short device code ("ime"); 3-6 digit serial we pad to 15.
  Col C  — client full name.
  Col D  — client phone.
  Col E  — channel name (Alif / Birzum / Hamroh / TBC / naqt / ...);
           may be combined: "Alif+birzum".
  Col F  — amount in UZS; may be combined: "5300000+6900000".
  Col G  — operator full name.
  Col H  — downpayment (kept as a note).
  Col P  — expense (kept as a note).
  Col Q  — comment (free text).

A second sheet, `nomerla`, holds Operator → personal/company phone mapping
(Isim / shaxsiy / kampaniya).

The importer is deliberately permissive: it skips rows it can't parse and
returns a structured result instead of raising. Sales already in the DB
that match (operator + sold_at-date + amount + imei) are skipped.
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import BinaryIO

from django.db import transaction
from django.utils import timezone

from apps.catalog.models import Channel
from apps.operators.models import Operator, OperatorStatus
from apps.sales.models import Sale, SaleOperator, SalePartner, SaleStatus

# Header words found in the first cell of non-sale rows — used to skip them.
_HEADER_TOKENS = {
    "ime",
    "mijoz",
    "hamkorlar",
    "ishchila",
    "daplata",
    "rasxodlar",
    "izoh",
    "mushtariy",
    "sevara",
    "raximjon",
    "ozodbek",
}

# Map noisy raw channel names to the canonical ones we want in the DB.
_CHANNEL_CANON = {
    "alif": "Alif",
    "uzum": "Uzum",
    "birzum": "Birzum",
    "hamroh": "Hamroh",
    "tbc": "TBC",
    "tbc birzum": "TBC",
    "naqt": "Cash",
    "walkin": "Walk-in",
    "walk-in": "Walk-in",
    "phone-call": "Phone-call",
    "phonecall": "Phone-call",
    "whatsapp": "WhatsApp",
}

_PLUS_SPLIT = re.compile(r"\s*\+\s*")


@dataclass
class ImportResult:
    sales_created: int = 0
    sales_skipped: int = 0
    operators_created: int = 0
    channels_created: int = 0
    errors: list[str] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "sales_created": self.sales_created,
            "sales_skipped": self.sales_skipped,
            "operators_created": self.operators_created,
            "channels_created": self.channels_created,
            "errors": self.errors[:50],
        }


def _canon_channel_name(raw: str) -> str:
    key = re.sub(r"\s+", " ", raw.strip().lower()).strip()
    key_nosp = key.replace(" ", "")
    return _CHANNEL_CANON.get(key) or _CHANNEL_CANON.get(key_nosp) or raw.strip().title()


def _parse_amounts(raw) -> list[Decimal]:
    """Return list of positive Decimals parsed from the cell."""
    if raw is None:
        return []
    if isinstance(raw, int | float | Decimal):
        try:
            v = Decimal(str(raw))
            return [v] if v > 0 else []
        except InvalidOperation:
            return []
    s = str(raw).strip()
    if not s:
        return []
    out: list[Decimal] = []
    for part in _PLUS_SPLIT.split(s):
        digits = re.sub(r"[^\d]", "", part)
        if digits:
            try:
                v = Decimal(digits)
                if v > 0:
                    out.append(v)
            except InvalidOperation:
                pass
    return out


def _parse_channels(raw) -> list[str]:
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    return [_canon_channel_name(p) for p in _PLUS_SPLIT.split(s) if p.strip()]


def _imei_from_code(code, *, fallback_seed: str) -> str:
    """
    Pad short device codes to 15 digits. When the code is missing, derive a
    deterministic 15-digit token from `fallback_seed` so re-importing the
    same row produces the same IMEI (and dedups correctly).
    """
    if isinstance(code, float) and code.is_integer():
        code = int(code)
    digits = re.sub(r"\D", "", "" if code is None else str(code))
    if digits:
        return digits[-15:].zfill(15)
    import hashlib

    h = int(hashlib.sha1(fallback_seed.encode("utf-8")).hexdigest(), 16)
    return f"{h % 10**15:015d}"


def _norm_operator_key(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip()).lower()


def _get_or_create_operator(name: str, phone: str = "", *, result: ImportResult) -> Operator:
    """Idempotent lookup by case-insensitive trimmed full_name."""
    display = re.sub(r"\s+", " ", name.strip()).title()
    if not display:
        display = "Unknown"
    op = Operator.objects.filter(full_name__iexact=display).first()
    if op:
        if phone and not op.phone:
            op.phone = phone
            op.save(update_fields=["phone", "updated_at"])
        return op
    op = Operator.objects.create(
        full_name=display, phone=phone or "", status=OperatorStatus.ACTIVE
    )
    result.operators_created += 1
    return op


def _get_or_create_channel(name: str, *, result: ImportResult) -> Channel:
    canon = _canon_channel_name(name)
    obj, created = Channel.objects.get_or_create(name=canon, defaults={"is_active": True})
    if created:
        result.channels_created += 1
    elif not obj.is_active:
        obj.is_active = True
        obj.save(update_fields=["is_active", "updated_at"])
    return obj


_DATE_STR_RE = re.compile(r"^\s*(\d{1,2})[.,/\-](\d{1,2})[.,/\-](\d{2,4})\s*$")


def _parse_date_string(s: str) -> dt.datetime | None:
    """Parse strings like '06,13,26' or '06.13.26' as MM,DD,YY (with DD,MM fallback)."""
    m = _DATE_STR_RE.match(s or "")
    if not m:
        return None
    a, b, c = (int(x) for x in m.groups())
    year = 2000 + c if c < 100 else c
    try:
        return dt.datetime(year, a, b)  # MM, DD
    except ValueError:
        try:
            return dt.datetime(year, b, a)  # DD, MM fallback
        except ValueError:
            return None


def _detect_dates_in_sheet(rows: list, tz) -> dict[int, dt.datetime]:
    """
    Walk rows once, collecting section-date markers and deciding whether
    openpyxl swallowed them as DD.MM.YY when the user meant MM.DD.YY.

    Heuristic: if every Excel-parsed datetime marker shares the same `day`
    and the `month` values vary, the user actually entered them as
    MM.DD.YY — swap month↔day. Confirmed when a sibling string-form marker
    like '06,13,26' exists that wouldn't parse as DD.MM (day 13 invalid).
    """
    parsed: list[tuple[int, dt.datetime]] = []
    strings: list[tuple[int, dt.datetime]] = []
    for i, row in enumerate(rows):
        if i == 0 or not row:
            continue
        a = row[0]
        if isinstance(a, dt.datetime):
            parsed.append((i, a))
        elif isinstance(a, str):
            d = _parse_date_string(a.strip())
            if d:
                strings.append((i, d))

    needs_swap = False
    if len(parsed) >= 2:
        days = {m.day for _, m in parsed}
        months = {m.month for _, m in parsed}
        if len(days) == 1 and len(months) > 1:
            needs_swap = True

    date_by_row: dict[int, dt.datetime] = {}
    for ri, d in parsed:
        if needs_swap:
            try:
                d = d.replace(month=d.day, day=d.month)
            except ValueError:
                pass
        date_by_row[ri] = d if d.tzinfo else d.replace(tzinfo=tz)
    for ri, d in strings:
        date_by_row[ri] = d if d.tzinfo else d.replace(tzinfo=tz)
    return date_by_row


def _is_header_row(cell_a: str) -> bool:
    s = cell_a.strip().lower()
    if not s:
        return True
    if any(s == tok or s.startswith(tok + " ") for tok in _HEADER_TOKENS):
        return True
    return False


def _import_operators_sheet(ws, result: ImportResult) -> None:
    """Sheet `nomerla`: Isim, shaxsiy, kampaniya."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = (row[0] or "").strip() if row and row[0] is not None else ""
        if not name or name.lower() in {"isim", "ism", "name"}:
            continue
        personal = row[1] if len(row) > 1 else None
        phone = ""
        if personal is not None:
            if isinstance(personal, float) and personal.is_integer():
                phone = str(int(personal))
            else:
                phone = re.sub(r"\s+", "", str(personal))
        _get_or_create_operator(name, phone, result=result)


def _import_sales_sheet(ws, result: ImportResult) -> None:
    tz = timezone.get_current_timezone()
    rows = list(ws.iter_rows(values_only=True))
    date_by_row = _detect_dates_in_sheet(rows, tz)

    current_date: dt.datetime | None = None
    for i, row in enumerate(rows):
        if i == 0 or not row:
            continue

        if i in date_by_row:
            current_date = date_by_row[i]
            continue

        a = row[0]
        cell_a = "" if a is None else str(a)
        if _is_header_row(cell_a):
            continue

        operator_raw = (row[6] or "").strip() if len(row) > 6 and row[6] else ""
        channel_raw = (row[4] or "") if len(row) > 4 else ""
        amount_raw = row[5] if len(row) > 5 else None

        amounts = _parse_amounts(amount_raw)
        if not amounts:
            # Some rows hold totals at the top — skip silently.
            continue
        if not operator_raw:
            continue

        operator = _get_or_create_operator(operator_raw, result=result)

        channels = _parse_channels(channel_raw)
        if not channels:
            channels = ["Other"]
        if len(channels) == len(amounts) and len(amounts) > 1:
            pairs = list(zip(channels, amounts, strict=False))
        else:
            # Pick first channel; if multiple amounts, sum them.
            pairs = [(channels[0], sum(amounts))]

        model_name = cell_a.strip()[:128] or "Не определена"
        client_name = (row[2] or "").strip() if len(row) > 2 and row[2] else ""
        client_phone = ""
        if len(row) > 3 and row[3] is not None:
            cp = row[3]
            if isinstance(cp, float) and cp.is_integer():
                client_phone = str(int(cp))
            else:
                client_phone = re.sub(r"\s+", "", str(cp))
        downpayment = row[7] if len(row) > 7 else None
        expense = row[15] if len(row) > 15 else None
        comment_cell = row[16] if len(row) > 16 else None

        comment_parts: list[str] = []
        if client_name:
            comment_parts.append(f"клиент: {client_name}")
        if client_phone:
            comment_parts.append(f"тел: {client_phone}")
        if downpayment:
            comment_parts.append(f"предоплата: {downpayment}")
        if expense:
            comment_parts.append(f"расход: {expense}")
        if comment_cell:
            comment_parts.append(str(comment_cell).strip())
        comment = " | ".join(comment_parts)[:2000]

        # Deterministic fallback so re-importing the same workbook is idempotent.
        sold_at = current_date or dt.datetime(1970, 1, 1, tzinfo=tz)
        total_amount = sum(amt for _, amt in pairs)
        primary_channel_name = pairs[0][0]
        primary_channel = _get_or_create_channel(primary_channel_name, result=result)
        imei = _imei_from_code(
            row[1] if len(row) > 1 else None,
            fallback_seed=(
                f"{model_name}|{operator.id}|{primary_channel_name}|"
                f"{total_amount}|{sold_at.isoformat()}"
            ),
        )

        # Idempotency: a sale uniquely identified by (operator, sold_at, total,
        # imei) — the same Excel row can't import twice.
        if Sale.objects.filter(
            operator=operator, sold_at=sold_at, amount=total_amount, imei=imei
        ).exists():
            result.sales_skipped += 1
            continue

        # One Sale per Excel row. Split payments become N partner_lines on
        # the same sale, not N separate sales (which would double-count
        # in KPIs / leaderboard / export).
        sale = Sale.objects.create(
            imei=imei,
            phone_model=model_name,
            operator=operator,
            channel=primary_channel,
            amount=total_amount,
            comment=comment,
            sold_at=sold_at,
            status=SaleStatus.CONFIRMED,
        )
        SaleOperator.objects.create(sale=sale, operator=operator, amount=total_amount)
        for ch_name, amt in pairs:
            ch = _get_or_create_channel(ch_name, result=result)
            SalePartner.objects.create(sale=sale, partner=ch, amount=amt)
        result.sales_created += 1


@transaction.atomic
def import_xlsx(source: str | BinaryIO, *, wipe_existing: bool = False) -> ImportResult:
    """Parse the `savdo`/`nomerla` workbook and load it into the DB."""
    import openpyxl

    wb = openpyxl.load_workbook(source, data_only=True, read_only=False)
    result = ImportResult()

    if wipe_existing:
        deleted = Sale.objects.all().delete()[0]
        result.errors.append(f"wiped {deleted} pre-existing sale(s)")

    # Sheet names sometimes carry trailing spaces — match case-insensitively.
    by_norm = {name.strip().lower(): name for name in wb.sheetnames}

    if "nomerla" in by_norm:
        _import_operators_sheet(wb[by_norm["nomerla"]], result)

    if "savdo" in by_norm:
        _import_sales_sheet(wb[by_norm["savdo"]], result)
    else:
        result.errors.append(
            f"worksheet 'savdo' not found (have: {list(wb.sheetnames)})"
        )

    return result

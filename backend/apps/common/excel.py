"""
Shared openpyxl helpers: every export in the project uses these so
formatting stays consistent across sheets.
"""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook as WB
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="111827")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTALS_FILL = PatternFill("solid", fgColor="F3F4F6")
TOTALS_FONT = Font(bold=True)
BORDER = Border(*[Side(style="thin", color="E5E7EB")] * 4)
MONEY_FMT = "#,##0_);[Red](#,##0);0"
INT_FMT = "#,##0"


def new_workbook() -> WB:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def write_sheet(
    wb: WB,
    *,
    title: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    money_columns: Sequence[int] = (),
    int_columns: Sequence[int] = (),
    totals_row: Sequence | None = None,
) -> Worksheet:
    """
    Build one sheet with bold header, optional totals row, money formatting,
    and auto-sized columns.

    `money_columns` and `int_columns` are 0-indexed positions.
    """
    ws = wb.create_sheet(title[:31])

    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for row in rows:
        ws.append(list(row))

    if totals_row is not None:
        ws.append(list(totals_row))
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.font = TOTALS_FONT
            cell.fill = TOTALS_FILL
            cell.border = BORDER

    for col_idx in money_columns:
        for cell in ws.iter_cols(min_col=col_idx + 1, max_col=col_idx + 1, min_row=2):
            for c in cell:
                c.number_format = MONEY_FMT
    for col_idx in int_columns:
        for cell in ws.iter_cols(min_col=col_idx + 1, max_col=col_idx + 1, min_row=2):
            for c in cell:
                c.number_format = INT_FMT

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for c in row:
                v = "" if c.value is None else str(c.value)
                if len(v) > max_len:
                    max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    return ws


def workbook_response(wb: WB, filename: str) -> HttpResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

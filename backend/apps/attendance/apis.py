import datetime as dt
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.throttling import AnonRateThrottle
from rest_framework.parsers import MultiPartParser, JSONParser, FormParser
from django.core.exceptions import PermissionDenied, ValidationError

from apps.operators.models import Operator
from apps.users.permissions import (
    _role,
    IsTeamLead,
    IsAuthenticatedAnyRole,
    IsSuperadminOrManager,
)
from apps.users.models import Role
from apps.common.pagination import DefaultPagination

from .models import AttendanceLog, AttendanceSettings
from .services import (
    attendance_scan,
    process_attendance_event,
    operator_qr_rotate,
    ScanRateLimitError,
    QrRevokedError,
    IpNotAllowedError,
    PhotoRequiredError,
    attendance_log_manual_close,
    attendance_log_backfill_checkout,
    qr_token_verify,
)
from .selectors import (
    attendance_dashboard_snapshot,
    attendance_settings_get,
    attendance_report,
    attendance_photos_queryset,
    logs_for_operator,
    open_log_for_operator,
    operator_qr_png_bytes,
    operator_qr_current,
    pending_backfill_log_for_operator,
)
from .permissions import IsTeamLeadOrManager, IsAttendancePinVerified


class AttendanceScanThrottle(AnonRateThrottle):
    scope = "attendance_scan_ip"


def _get_client_ip(request) -> str:
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        ip = x_forwarded_for.split(",")[0].strip()
    else:
        ip = request.META.get("REMOTE_ADDR")
    return ip


def _extract_photo_bytes(request) -> tuple[bytes | None, str | None]:
    """Extract raw bytes + filename from multipart `photo` upload."""
    photo = request.FILES.get("photo") if hasattr(request, "FILES") else None
    if not photo:
        return None, None
    return photo.read(), (photo.name or None)


class ScanAttendanceApi(APIView):
    permission_classes = [AllowAny]
    throttle_classes = [AttendanceScanThrottle]
    # Accept JSON (legacy: {qr_payload}) OR multipart (adds `photo` file).
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def post(self, request):
        qr_payload = request.data.get("qr_payload")
        if not qr_payload:
            return Response(
                {"error": "qr_payload is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ip = _get_client_ip(request)
        user_agent = request.META.get("HTTP_USER_AGENT", "")[:256]
        photo_bytes, photo_filename = _extract_photo_bytes(request)

        try:
            res = attendance_scan(
                qr_raw=qr_payload,
                ip=ip,
                user_agent=user_agent,
                photo_bytes=photo_bytes,
                photo_filename=photo_filename,
            )
            return Response(res, status=status.HTTP_200_OK)
        except ScanRateLimitError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_429_TOO_MANY_REQUESTS)
        except QrRevokedError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_410_GONE)
        except IpNotAllowedError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_403_FORBIDDEN)
        except PhotoRequiredError as exc:
            return Response({"error": str(exc), "code": "photo_required"}, status=status.HTTP_400_BAD_REQUEST)
        except ValidationError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)


class ScanWithPhotoAttendanceApi(APIView):
    """
    Mobile self-check-in with mandatory photo. Public (uses HMAC-signed QR
    payload for identity) — mirrors ScanAttendanceApi semantics but always
    treats photo as required + face-checked.
    """

    permission_classes = [AllowAny]
    throttle_classes = [AttendanceScanThrottle]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        qr_payload = request.data.get("qr_payload")
        if not qr_payload:
            return Response(
                {"error": "qr_payload is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        photo_bytes, photo_filename = _extract_photo_bytes(request)
        if not photo_bytes:
            return Response(
                {"error": "photo is required", "code": "photo_required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ip = _get_client_ip(request)
        user_agent = request.META.get("HTTP_USER_AGENT", "")[:256]

        try:
            operator, _qr = qr_token_verify(qr_payload)
        except QrRevokedError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_410_GONE)
        except ValidationError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            res = process_attendance_event(
                operator=operator,
                source="qr",
                initiator=f"scan-with-photo ip={ip or '-'}",
                ip=ip,
                user_agent=user_agent,
                issue_token=True,
                photo_bytes=photo_bytes,
                photo_filename=photo_filename,
                require_photo_override=True,
            )
            return Response(res, status=status.HTTP_200_OK)
        except ScanRateLimitError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_429_TOO_MANY_REQUESTS)
        except IpNotAllowedError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_403_FORBIDDEN)
        except PhotoRequiredError as exc:
            return Response({"error": str(exc), "code": "photo_required"}, status=status.HTTP_400_BAD_REQUEST)
        except ValidationError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)


class QrPreviewAttendanceApi(APIView):
    """Return operator name / status from QR without side-effects (no check-in).

    Lets the scan photo screen greet the operator by name before they snap.
    Public because the QR itself is the credential.
    """

    permission_classes = [AllowAny]

    def get(self, request):
        qr_payload = request.query_params.get("qr", "").strip()
        if not qr_payload:
            return Response({"error": "qr required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            operator, _ = qr_token_verify(qr_payload)
        except QrRevokedError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_410_GONE)
        except ValidationError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        from .selectors import open_log_for_operator

        open_log = open_log_for_operator(operator)
        return Response(
            {
                "operator": {"id": operator.id, "full_name": operator.full_name},
                "on_shift": open_log is not None,
                "checked_in_at": open_log.checked_in_at.isoformat() if open_log else None,
                "expected_action": "check_out" if open_log else "check_in",
            }
        )


class MeToggleAttendanceApi(APIView):
    """
    Auth-based check-in/out for the /profile card + operator dashboard
    widget. Accepts optional multipart `photo` — if provided it goes
    through the same validation as scan-with-photo.
    """

    permission_classes = [IsAuthenticated, IsAuthenticatedAnyRole]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def post(self, request):
        profile = getattr(request.user, "profile", None)
        if not profile or not profile.operator_id:
            return Response(
                {"error": "No operator linked to this user"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        ip = _get_client_ip(request)
        ua = request.META.get("HTTP_USER_AGENT", "")[:256]
        photo_bytes, photo_filename = _extract_photo_bytes(request)
        # `require_photo=1` from the frontend widget forces the photo gate
        # for that specific request even when the global setting is off —
        # keeps the "always ask for photo from the widget" UX consistent.
        require_photo_override = str(request.data.get("require_photo", "")).lower() in {
            "1",
            "true",
            "yes",
        }
        try:
            res = process_attendance_event(
                operator=profile.operator,
                source="manual",
                initiator=f"profile:{request.user.username}",
                ip=ip,
                user_agent=ua,
                issue_token=False,
                photo_bytes=photo_bytes,
                photo_filename=photo_filename,
                require_photo_override=require_photo_override,
            )
            return Response(res, status=status.HTTP_200_OK)
        except ScanRateLimitError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_429_TOO_MANY_REQUESTS)
        except PhotoRequiredError as exc:
            return Response({"error": str(exc), "code": "photo_required"}, status=status.HTTP_400_BAD_REQUEST)
        except ValidationError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)


class MeCurrentAttendanceApi(APIView):
    permission_classes = [IsAuthenticated, IsAuthenticatedAnyRole]

    def get(self, request):
        profile = getattr(request.user, "profile", None)
        if not profile or not profile.operator_id:
            return Response(
                {
                    "open_log": None,
                    "today_events": [],
                    "require_checkin_enabled": False,
                    "checkout_reminder_active": False,
                    "pending_backfill_log": None,
                },
                status=status.HTTP_200_OK,
            )

        operator = profile.operator
        open_log = open_log_for_operator(operator)

        # Today's events (Tashkent calendar day)
        today = timezone.localdate()
        logs_today = logs_for_operator(operator, since=today, until=today)

        today_events = []
        for l in logs_today:
            today_events.append({
                "id": l.id,
                "checked_in_at": l.checked_in_at.isoformat(),
                "checked_out_at": l.checked_out_at.isoformat() if l.checked_out_at else None,
                "was_late": l.was_late,
                "auto_closed": l.auto_closed,
            })

        open_log_data = None
        if open_log:
            open_log_data = {
                "id": open_log.id,
                "checked_in_at": open_log.checked_in_at.isoformat(),
                "was_late": open_log.was_late,
                "checkin_photo_url": open_log.checkin_photo.url if open_log.checkin_photo else None,
                "checkout_reminder_sent_at": (
                    open_log.checkout_reminder_sent_at.isoformat()
                    if open_log.checkout_reminder_sent_at
                    else None
                ),
            }

        # Enforcement wave 2026-08-26.
        # `require_checkin_enabled` — фронт использует, чтобы решить, показывать
        # ли fullscreen-gate. Backend НИ ОДИН endpoint не блокирует по этому
        # флагу — enforcement только на UI.
        require_checkin = bool(operator.require_checkin_enabled)

        # Reminder активен, если cron уже пометил open_log — фронт покажет
        # оранжевый баннер, локально приглушает через localStorage-dismiss.
        checkout_reminder_active = bool(
            open_log and open_log.checkout_reminder_sent_at is not None
        )

        # «Вчера забыли выйти» — есть auto_closed лог без backfill за
        # последние 3 дня. Фронт показывает блокирующий модал с time-picker.
        #
        # ВАЖНО (prod-safety, 2026-08-26): pending_backfill возвращаем ТОЛЬКО
        # операторам с `require_checkin_enabled=True`. Иначе prod-операторы,
        # которые исторически не отмечались, получили бы кучу старых
        # auto_closed логов и утром словили бы блокирующий backfill-модал —
        # регрессия ломала бы их работу. Enforcement включается per-operator.
        pending_backfill_data = None
        if require_checkin:
            pending_backfill = pending_backfill_log_for_operator(operator)
            if pending_backfill is not None:
                pending_backfill_data = {
                    "id": pending_backfill.id,
                    "checked_in_at": pending_backfill.checked_in_at.isoformat(),
                    "auto_closed_at": (
                        pending_backfill.checked_out_at.isoformat()
                        if pending_backfill.checked_out_at
                        else None
                    ),
                }

        return Response(
            {
                "open_log": open_log_data,
                "today_events": today_events,
                "operator_id": profile.operator_id,
                "require_checkin_enabled": require_checkin,
                "checkout_reminder_active": checkout_reminder_active,
                "pending_backfill_log": pending_backfill_data,
            },
            status=status.HTTP_200_OK,
        )


class MeBackfillCheckoutAttendanceApi(APIView):
    """
    Оператор задним числом вводит фактическое время своего ухода, когда
    вчера забыл /checkout и cron auto-close'нул смену в 23:00.

    Body: `{ log_id: int, checked_out_at: "YYYY-MM-DDTHH:MM" (local naive)
    или ISO с tz }`.

    Валидация в service (`attendance_log_backfill_checkout`):
      - лог принадлежит текущему оператору,
      - лог был `auto_closed=True` и ещё не backfilled,
      - `checked_out_at` в окне [check_in + 30мин, check_in + N часов],
      - `checked_out_at <= log.checked_out_at` (момент cron auto-close).

    Возвращает обновлённый лог. Фронт снимает блокирующий модал по
    успешному ответу и refetch'ит `/attendance/me/current/`.
    """

    permission_classes = [IsAuthenticated, IsAuthenticatedAnyRole]

    def post(self, request):
        profile = getattr(request.user, "profile", None)
        if not profile or not profile.operator_id:
            return Response(
                {"error": "Нет привязанного оператора"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        log_id_raw = request.data.get("log_id")
        checked_out_raw = request.data.get("checked_out_at")

        try:
            log_id = int(log_id_raw)
        except (TypeError, ValueError):
            return Response(
                {"error": "log_id обязателен и должен быть числом"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not checked_out_raw or not isinstance(checked_out_raw, str):
            return Response(
                {"error": "checked_out_at обязателен (ISO datetime)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Принимаем ISO (`2026-08-25T18:30:00`) без tz — считаем локальным
        # временем Ташкента. С tz — уважаем tz. `parse_datetime` вернёт
        # None если строка не парсится.
        from django.utils.dateparse import parse_datetime

        checked_out_at = parse_datetime(checked_out_raw)
        if checked_out_at is None:
            return Response(
                {"error": "Некорректный формат checked_out_at (ожидается ISO datetime)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            log = AttendanceLog.objects.get(id=log_id)
        except AttendanceLog.DoesNotExist:
            return Response({"error": "Лог не найден"}, status=status.HTTP_404_NOT_FOUND)

        try:
            log = attendance_log_backfill_checkout(
                operator=profile.operator,
                log=log,
                checked_out_at=checked_out_at,
            )
        except PermissionDenied as exc:
            return Response({"error": str(exc)}, status=status.HTTP_403_FORBIDDEN)
        except ValidationError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {
                "id": log.id,
                "operator_id": log.operator_id,
                "checked_in_at": log.checked_in_at.isoformat(),
                "checked_out_at": log.checked_out_at.isoformat(),
                "auto_closed": log.auto_closed,
                "backfilled_by_operator_at": log.backfilled_by_operator_at.isoformat(),
                "duration_min": log.duration_seconds // 60
                if log.duration_seconds is not None
                else None,
            },
            status=status.HTTP_200_OK,
        )


class MeHistoryAttendanceApi(APIView):
    permission_classes = [IsAuthenticated, IsAuthenticatedAnyRole]

    def get(self, request):
        profile = getattr(request.user, "profile", None)
        if not profile or not profile.operator_id:
            return Response({"error": "No operator profile link"}, status=status.HTTP_400_BAD_REQUEST)

        # parse from/to
        from_str = request.query_params.get("from")
        to_str = request.query_params.get("to")

        today = timezone.localdate()
        since = (
            dt.datetime.strptime(from_str, "%Y-%m-%d").date()
            if from_str
            else today - dt.timedelta(days=30)
        )
        until = dt.datetime.strptime(to_str, "%Y-%m-%d").date() if to_str else today

        logs = logs_for_operator(profile.operator, since=since, until=until)

        paginator = DefaultPagination()
        paginated_logs = paginator.paginate_queryset(logs, request)

        data = []
        for l in paginated_logs:
            data.append({
                "id": l.id,
                "checked_in_at": l.checked_in_at.isoformat(),
                "checked_out_at": l.checked_out_at.isoformat() if l.checked_out_at else None,
                "was_late": l.was_late,
                "auto_closed": l.auto_closed,
                "manually_closed": l.manually_closed,
                "manually_closed_by_name": l.manually_closed_by.username
                if l.manually_closed_by
                else None,
                "manual_close_note": l.manual_close_note,
                "duration_min": l.duration_seconds // 60 if l.duration_seconds is not None else None,
                "source": l.source,
            })

        return paginator.get_paginated_response(data)


class MeQrAttendanceApi(APIView):
    permission_classes = [IsAuthenticated, IsAuthenticatedAnyRole]

    def get(self, request):
        profile = getattr(request.user, "profile", None)
        if not profile or not profile.operator_id:
            return Response(
                {"error": "Пользователь не привязан к оператору"},
                status=status.HTTP_404_NOT_FOUND,
            )

        origin_hint = request.query_params.get("origin", "").strip() or None
        png_bytes = operator_qr_png_bytes(
            profile.operator, request=request, origin_hint=origin_hint
        )
        response = HttpResponse(png_bytes, content_type="image/png")
        response["Cache-Control"] = "private, no-store"
        response["Content-Disposition"] = "inline; filename=qr.png"
        return response


class MeQrTokenAttendanceApi(APIView):
    """
    Return the raw HMAC-signed QR payload + scannable URL for the *current*
    authenticated operator. Mirrors `OperatorQrTokenAttendanceApi` but is
    scoped to `request.user.profile.operator` — every operator can only
    fetch their own token, no cross-operator lookup.

    Used by the operator dashboard `AttendanceStatusWidget`: after tapping
    «Начать смену» / «Завершить смену» the widget opens a modal with a
    big QR rendered client-side from `url`. Operator scans it with their
    phone camera → phone opens `/scan?qr=<payload>` → ScanPhotoFlow
    handles the selfie + submit (no login required).

    Rationale: desktop PCs at the shop lack webcams; forcing the operator
    to snap on the same machine is impossible. Delegating photo capture
    to the personal phone via QR is the only ergonomic option.
    """

    permission_classes = [IsAuthenticated, IsAuthenticatedAnyRole]

    def get(self, request):
        from .selectors import operator_qr_current_or_create, build_scan_url
        from .services import qr_token_build

        profile = getattr(request.user, "profile", None)
        if not profile or not profile.operator_id:
            return Response(
                {"error": "Пользователь не привязан к оператору"},
                status=status.HTTP_404_NOT_FOUND,
            )

        operator = profile.operator
        qr_obj = operator_qr_current_or_create(operator)
        payload = qr_token_build(operator, qr_obj.nonce)

        origin_hint = request.query_params.get("origin", "").strip() or None
        url = build_scan_url(payload, request=request, origin_hint=origin_hint)

        return Response(
            {
                "operator_id": operator.id,
                "operator_name": operator.full_name,
                "payload": payload,
                "url": url,
                "nonce_prefix": qr_obj.nonce[:6],
            },
            status=status.HTTP_200_OK,
        )


class ReportAttendanceApi(APIView):
    permission_classes = [IsAuthenticated, IsTeamLeadOrManager, IsAttendancePinVerified]

    def get(self, request):
        date_from_str = request.query_params.get("date_from")
        date_to_str = request.query_params.get("date_to")
        operator_id_str = request.query_params.get("operator")
        fmt = request.query_params.get("format", "json")

        if date_from_str and date_to_str:
            try:
                date_from = dt.datetime.strptime(date_from_str, "%Y-%m-%d").date()
                date_to = dt.datetime.strptime(date_to_str, "%Y-%m-%d").date()
            except ValueError:
                return Response(
                    {"error": "Invalid date format, use YYYY-MM-DD"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            operator_ids = None
            if operator_id_str:
                operator_ids = [int(x) for x in operator_id_str.split(",") if x.strip()]

            from .selectors import attendance_statistics_report

            report = attendance_statistics_report(
                date_from=date_from, date_to=date_to, operator_ids=operator_ids
            )

            if fmt == "xlsx":
                from apps.common.excel import (
                    new_workbook,
                    write_sheet,
                    workbook_response,
                    BORDER,
                    HEADER_FONT,
                    HEADER_FILL,
                )
                from openpyxl.styles import Alignment
                from openpyxl.utils import get_column_letter

                wb = new_workbook()
                headers = [
                    "Оператор",
                    "Присутствовал",
                    "Ожидалось",
                    "Пропустил",
                    "Опоздал",
                    "Среднее опоздание (мин)",
                    "Авто-закрыто",
                    "Ручное закрытие",
                    "Средняя длина смены (мин)",
                    "Итого часов",
                ]
                rows = []
                for row in report["rows"]:
                    rows.append([
                        row["operator_name"],
                        row["days_present"],
                        row["days_expected"],
                        row["days_absent"],
                        row["late_count"],
                        row["avg_late_minutes"],
                        row["auto_closed_count"],
                        row["manually_closed_count"],
                        row["avg_shift_minutes"],
                        row["total_worked_hours"],
                    ])

                write_sheet(wb, title="Attendance", headers=headers, rows=rows)

                ws2 = wb.create_sheet("Heatmap")
                dates = [x["date"] for x in report["rows"][0]["heatmap"]] if report["rows"] else []
                ws2.append(["Оператор"] + dates)

                for cell in ws2[1]:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.border = BORDER
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for row in report["rows"]:
                    op_name = row["operator_name"]
                    statuses = [x["status"] for x in row["heatmap"]]
                    ws2.append([op_name] + statuses)

                for col in ws2.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws2.column_dimensions[col_letter].width = min(max_len + 2, 60)

                return workbook_response(wb, f"attendance_{date_from_str}_to_{date_to_str}.xlsx")

            return Response(report, status=status.HTTP_200_OK)

        date_str = request.query_params.get("date")
        if not date_str or date_str == "today":
            day = timezone.localdate()
        else:
            try:
                day = dt.datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                return Response(
                    {"error": "Invalid date format, use YYYY-MM-DD"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        report = attendance_report(day)
        return Response(report, status=status.HTTP_200_OK)



class OperatorLogsAttendanceApi(APIView):
    permission_classes = [IsAuthenticated, IsTeamLeadOrManager, IsAttendancePinVerified]

    def get(self, request, operator_id):
        operator = get_object_or_404(Operator, id=operator_id)

        from_str = request.query_params.get("from")
        to_str = request.query_params.get("to")

        today = timezone.localdate()
        since = (
            dt.datetime.strptime(from_str, "%Y-%m-%d").date()
            if from_str
            else today - dt.timedelta(days=30)
        )
        until = dt.datetime.strptime(to_str, "%Y-%m-%d").date() if to_str else today

        logs = logs_for_operator(operator, since=since, until=until)

        data = []
        for l in logs:
            data.append({
                "id": l.id,
                "checked_in_at": l.checked_in_at.isoformat(),
                "checked_out_at": l.checked_out_at.isoformat() if l.checked_out_at else None,
                "was_late": l.was_late,
                "auto_closed": l.auto_closed,
                "manually_closed": l.manually_closed,
                "manually_closed_by_name": l.manually_closed_by.username
                if l.manually_closed_by
                else None,
                "manual_close_note": l.manual_close_note,
                "duration_min": l.duration_seconds // 60 if l.duration_seconds is not None else None,
                "source": l.source,
            })

        return Response(data, status=status.HTTP_200_OK)


class OperatorQrAttendanceApi(APIView):
    permission_classes = [IsAuthenticated, IsTeamLeadOrManager, IsAttendancePinVerified]

    def get(self, request, operator_id):
        operator = get_object_or_404(Operator, id=operator_id)
        qr = operator_qr_current(operator)

        if not qr:
            return Response(
                {"nonce": None, "created_at": None, "revoked_at": None},
                status=status.HTTP_200_OK,
            )

        return Response(
            {
                "nonce": qr.nonce[:6] + "...",
                "created_at": qr.created_at.isoformat(),
                "revoked_at": qr.revoked_at.isoformat() if qr.revoked_at else None,
            },
            status=status.HTTP_200_OK,
        )


class OperatorQrTokenAttendanceApi(APIView):
    """
    Return the raw HMAC-signed QR payload + a scannable URL for a given
    operator. Used by the manager Kiosk page:

    - `payload` — the token string that fits `POST /scan-with-photo/` as
      `qr_payload` when the manager captures a photo directly from a
      laptop/desktop webcam (no round-trip through the phone).
    - `url`     — the same payload wrapped as
      `${QR_CHECKIN_URL}?qr=<payload>`, ready to render as a QR that the
      operator scans with their phone camera — opens `/scan?qr=` on the
      phone, which then does check-in without login (public HMAC flow).

    Manager-gated (superadmin included via SENIOR_ROLES). This endpoint
    is intentionally not exposed to operators — they use their own
    `GET /api/me/attendance-qr.png` and `GET /api/attendance/qr-preview/`
    for the same purpose.
    """

    permission_classes = [IsAuthenticated, IsTeamLeadOrManager, IsAttendancePinVerified]

    def get(self, request, operator_id):
        from .selectors import operator_qr_current_or_create, build_scan_url
        from .services import qr_token_build

        operator = get_object_or_404(Operator, id=operator_id)
        qr_obj = operator_qr_current_or_create(operator)
        payload = qr_token_build(operator, qr_obj.nonce)

        origin_hint = request.query_params.get("origin", "").strip() or None
        url = build_scan_url(payload, request=request, origin_hint=origin_hint)

        return Response(
            {
                "operator_id": operator.id,
                "operator_name": operator.full_name,
                "payload": payload,
                "url": url,
                "nonce_prefix": qr_obj.nonce[:6],
            },
            status=status.HTTP_200_OK,
        )


class OperatorQrRotateAttendanceApi(APIView):
    permission_classes = [IsAuthenticated, IsTeamLead, IsAttendancePinVerified]

    def post(self, request, operator_id):
        operator = get_object_or_404(Operator, id=operator_id)
        qr = operator_qr_rotate(operator=operator, actor=request.user)

        return Response(
            {
                "nonce": qr.nonce[:6] + "...",
                "created_at": qr.created_at.isoformat(),
                "png_url": f"/api/attendance/operators/{operator.id}/qr.png",
            },
            status=status.HTTP_200_OK,
        )


class OperatorQrPngAttendanceApi(APIView):
    permission_classes = [IsAuthenticated, IsTeamLeadOrManager, IsAttendancePinVerified]

    def get(self, request, operator_id):
        operator = get_object_or_404(Operator, id=operator_id)
        origin_hint = request.query_params.get("origin", "").strip() or None
        png_bytes = operator_qr_png_bytes(
            operator, request=request, origin_hint=origin_hint
        )

        response = HttpResponse(png_bytes, content_type="image/png")
        response["Cache-Control"] = "private, no-store"
        response["Content-Disposition"] = f"inline; filename=qr-{operator_id}.png"
        return response


class SettingsAttendanceApi(APIView):
    permission_classes = [IsAuthenticated, IsTeamLead, IsAttendancePinVerified]

    def _serialize(self, settings_obj) -> dict:
        return {
            "shift_start": settings_obj.shift_start.strftime("%H:%M")
            if isinstance(settings_obj.shift_start, dt.time)
            else settings_obj.shift_start,
            "shift_end": settings_obj.shift_end.strftime("%H:%M")
            if isinstance(settings_obj.shift_end, dt.time)
            else settings_obj.shift_end,
            "late_threshold_min": settings_obj.late_threshold_min,
            "auto_close_at": settings_obj.auto_close_at.strftime("%H:%M")
            if isinstance(settings_obj.auto_close_at, dt.time)
            else settings_obj.auto_close_at,
            "tg_checkin_enabled": settings_obj.tg_checkin_enabled,
            "require_photo": settings_obj.require_photo,
            "require_face": settings_obj.require_face,
            "photo_max_size_mb": settings_obj.photo_max_size_mb,
            "checkout_reminder_after_hours": settings_obj.checkout_reminder_after_hours,
            "max_backfill_hours": settings_obj.max_backfill_hours,
            # Payroll-defaults.
            "default_salary_uzs": str(settings_obj.default_salary_uzs),
            "default_grace_period_min": settings_obj.default_grace_period_min,
            "default_late_penalty_uzs": str(settings_obj.default_late_penalty_uzs),
            "default_weekly_day_off": settings_obj.default_weekly_day_off,
            "default_attendance_gate_pct": settings_obj.default_attendance_gate_pct,
            "default_weekly_free_absences": settings_obj.default_weekly_free_absences,
        }

    def get(self, request):
        return Response(self._serialize(attendance_settings_get()))

    def patch(self, request):
        settings_obj = attendance_settings_get()

        shift_start = request.data.get("shift_start")
        shift_end = request.data.get("shift_end")
        late_threshold_min = request.data.get("late_threshold_min")
        auto_close_at = request.data.get("auto_close_at")
        tg_checkin_enabled = request.data.get("tg_checkin_enabled")
        require_photo = request.data.get("require_photo")
        require_face = request.data.get("require_face")
        photo_max_size_mb = request.data.get("photo_max_size_mb")

        if shift_start:
            settings_obj.shift_start = shift_start
        if shift_end:
            settings_obj.shift_end = shift_end
        if late_threshold_min is not None:
            settings_obj.late_threshold_min = int(late_threshold_min)
        if auto_close_at:
            settings_obj.auto_close_at = auto_close_at
        if tg_checkin_enabled is not None:
            settings_obj.tg_checkin_enabled = bool(tg_checkin_enabled)
        if require_photo is not None:
            settings_obj.require_photo = bool(require_photo)
        if require_face is not None:
            settings_obj.require_face = bool(require_face)
        if photo_max_size_mb is not None:
            settings_obj.photo_max_size_mb = max(1, min(20, int(photo_max_size_mb)))
        checkout_reminder_after_hours = request.data.get("checkout_reminder_after_hours")
        max_backfill_hours = request.data.get("max_backfill_hours")
        if checkout_reminder_after_hours is not None:
            settings_obj.checkout_reminder_after_hours = max(
                0, min(24, int(checkout_reminder_after_hours))
            )
        if max_backfill_hours is not None:
            settings_obj.max_backfill_hours = max(1, min(24, int(max_backfill_hours)))

        # Payroll defaults.
        from decimal import Decimal as _D

        default_salary_uzs = request.data.get("default_salary_uzs")
        default_grace_period_min = request.data.get("default_grace_period_min")
        default_late_penalty_uzs = request.data.get("default_late_penalty_uzs")
        default_weekly_day_off = request.data.get("default_weekly_day_off")
        default_attendance_gate_pct = request.data.get("default_attendance_gate_pct")
        default_weekly_free_absences = request.data.get("default_weekly_free_absences")
        if default_salary_uzs is not None:
            settings_obj.default_salary_uzs = _D(str(default_salary_uzs))
        if default_grace_period_min is not None:
            settings_obj.default_grace_period_min = max(
                0, min(240, int(default_grace_period_min))
            )
        if default_late_penalty_uzs is not None:
            settings_obj.default_late_penalty_uzs = _D(str(default_late_penalty_uzs))
        if default_weekly_day_off is not None:
            settings_obj.default_weekly_day_off = max(
                0, min(6, int(default_weekly_day_off))
            )
        if default_attendance_gate_pct is not None:
            settings_obj.default_attendance_gate_pct = max(
                0, min(100, int(default_attendance_gate_pct))
            )
        if default_weekly_free_absences is not None:
            settings_obj.default_weekly_free_absences = max(
                0, min(7, int(default_weekly_free_absences))
            )

        settings_obj.updated_by = request.user
        settings_obj.save()
        return Response(self._serialize(settings_obj))


class PhotosGalleryAttendanceApi(APIView):
    """
    GET /api/attendance/photos/ — постраничная лента check-in / check-out
    фото по всем операторам.

    Query params:
      - `date_from` / `date_to` — YYYY-MM-DD (опц., по календарю Ташкента);
      - `operator` (или `operator_id`) — фильтр по одному оператору;
      - `limit` / `offset` — стандартная DRF-пагинация (DefaultPagination).

    Доступ: `IsSuperadminOrManager` — оба senior-роли + superadmin. Оператор
    получит 403.
    """

    permission_classes = [IsAuthenticated, IsSuperadminOrManager, IsAttendancePinVerified]

    def get(self, request):
        # Parse dates
        date_from_str = request.query_params.get("date_from") or None
        date_to_str = request.query_params.get("date_to") or None
        operator_id_str = (
            request.query_params.get("operator")
            or request.query_params.get("operator_id")
            or None
        )

        date_from = None
        date_to = None
        if date_from_str:
            try:
                date_from = dt.datetime.strptime(date_from_str, "%Y-%m-%d").date()
            except ValueError:
                return Response(
                    {"error": "date_from: YYYY-MM-DD"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        if date_to_str:
            try:
                date_to = dt.datetime.strptime(date_to_str, "%Y-%m-%d").date()
            except ValueError:
                return Response(
                    {"error": "date_to: YYYY-MM-DD"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        operator_id: int | None = None
        if operator_id_str:
            try:
                operator_id = int(operator_id_str)
            except (TypeError, ValueError):
                return Response(
                    {"error": "operator must be int"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        qs = attendance_photos_queryset(
            date_from=date_from, date_to=date_to, operator_id=operator_id
        )

        paginator = DefaultPagination()
        page = paginator.paginate_queryset(qs, request)

        data = []
        for log in page:
            local_in = timezone.localtime(log.checked_in_at)
            data.append({
                "log_id": log.id,
                "operator_id": log.operator_id,
                "operator_name": log.operator.full_name,
                "date": local_in.strftime("%Y-%m-%d"),
                "checked_in_at": log.checked_in_at.isoformat(),
                "checked_out_at": log.checked_out_at.isoformat()
                if log.checked_out_at
                else None,
                "checkin_photo_url": log.checkin_photo.url if log.checkin_photo else None,
                # Тумбы пока = сам URL (клиент рендерит <img> с CSS-thumbом).
                # При появлении отдельного thumb-варианта достаточно поменять
                # только это поле — контракт клиента не сломается.
                "checkin_photo_thumb": log.checkin_photo.url if log.checkin_photo else None,
                "checkout_photo_url": log.checkout_photo.url if log.checkout_photo else None,
                "checkout_photo_thumb": log.checkout_photo.url if log.checkout_photo else None,
                "was_late": log.was_late,
                "auto_closed": log.auto_closed,
                "manually_closed": log.manually_closed,
                "source": log.source,
                "duration_min": log.duration_seconds // 60
                if log.duration_seconds is not None
                else None,
            })

        return paginator.get_paginated_response(data)


class ManualCloseAttendanceApi(APIView):
    permission_classes = [IsAuthenticated, IsTeamLeadOrManager, IsAttendancePinVerified]

    def post(self, request, log_id):
        log = get_object_or_404(AttendanceLog, id=log_id)
        note = request.data.get("note", "").strip()
        if len(note) > 280:
            return Response(
                {"error": "Note is too long (max 280 characters)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if log.checked_out_at is not None:
            return Response(
                {"detail": "Лог уже закрыт"},
                status=status.HTTP_409_CONFLICT,
            )

        try:
            log = attendance_log_manual_close(log=log, user=request.user, note=note)
        except ValidationError as exc:
            return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {
                "id": log.id,
                "operator_id": log.operator_id,
                "checked_in_at": log.checked_in_at.isoformat(),
                "checked_out_at": log.checked_out_at.isoformat(),
                "was_late": log.was_late,
                "auto_closed": log.auto_closed,
                "manually_closed": log.manually_closed,
                "manually_closed_by": log.manually_closed_by_id,
                "manual_close_note": log.manual_close_note,
                "duration_min": log.duration_seconds // 60
                if log.duration_seconds is not None
                else None,
                "source": log.source,
            },
            status=status.HTTP_200_OK,
        )


def _parse_month_param(month_str: str | None) -> tuple[int, int] | None:
    """
    Парсит `?month=YYYY-MM` в `(year, month)`. Возвращает None если
    отсутствует / битый формат — вызывающий сам решает fallback
    (обычно = текущий месяц Ташкента).
    """
    if not month_str:
        return None
    try:
        parts = month_str.split("-")
        if len(parts) != 2:
            return None
        year = int(parts[0])
        month = int(parts[1])
        if not (1 <= month <= 12) or year < 2000 or year > 3000:
            return None
        return year, month
    except (ValueError, TypeError):
        return None


def _default_month() -> tuple[int, int]:
    today = timezone.localdate()
    return today.year, today.month


def _payroll_summary_to_xlsx(summary: dict, filename: str):
    """
    Excel: две вкладки — «Сводка» и «По дням». Форматирование денег /
    заголовков — через shared helper'ы `apps.common.excel`.
    """
    from apps.common.excel import new_workbook, workbook_response, write_sheet

    wb = new_workbook()

    # Sheet 1: Сводка (2 колонки key/value).
    summary_headers = ["Показатель", "Значение"]
    summary_rows = [
        ["Оператор", summary.get("operator_name", "")],
        ["Период", f"{summary.get('year')}-{int(summary.get('month', 1)):02d}"],
        ["Оклад (UZS)", int(summary.get("salary_gross", 0) or 0)],
        ["Рабочих дней (план)", int(summary.get("working_days_planned", 0))],
        ["Посещал", int(summary.get("days_attended", 0))],
        ["Пропустил (всего)", int(summary.get("days_absent", 0))],
        ["Прощённых пропусков", int(summary.get("weekly_free_absences_used", 0))],
        ["Штрафуемых пропусков", int(summary.get("billable_absences", 0))],
        ["Опозданий", int(summary.get("days_late", 0))],
        ["Ср. опоздание (мин)", int(summary.get("avg_late_minutes", 0))],
        ["Посещаемость (%)", float(summary.get("attendance_rate_pct", 0))],
        ["Гейт (%)", int(summary.get("gate_pct", 0))],
        [
            "Гейт сработал",
            "Да (оклад = 0)" if summary.get("gate_triggered") else "Нет",
        ],
        ["Дневная ставка (UZS)", int(summary.get("daily_rate", 0) or 0)],
        [
            "Вычет за отсутствия (UZS)",
            int(summary.get("absence_deduction", 0) or 0),
        ],
        [
            "Штраф за опоздание (UZS/шт)",
            int(summary.get("late_penalty_per_event", 0) or 0),
        ],
        [
            "Штраф за опоздания (UZS)",
            int(summary.get("late_penalty_total", 0) or 0),
        ],
        ["К выплате (UZS)", int(summary.get("salary_earned", 0) or 0)],
    ]
    # Колонки «Значение» — money-fmt только на числовых Rows. Т.к.
    # write_sheet применяет формат к целой колонке кроме заголовка, а у
    # нас смешан текст/число — оставляем формат `INT_FMT` глобально, а
    # текстовые cell'ы просто отобразятся как строки.
    write_sheet(
        wb,
        title="Сводка",
        headers=summary_headers,
        rows=summary_rows,
        int_columns=(1,),
    )

    # Sheet 2: По дням.
    day_headers = [
        "Дата",
        "День недели",
        "Рабочий",
        "Статус",
        "Приход",
        "Уход",
        "Опоздание (мин)",
        "Вычет (UZS)",
        "Комментарий",
    ]
    weekday_ru = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    status_ru = {
        "on_time": "В срок",
        "late": "Опоздал",
        "absent": "Отсутствовал",
        "free_absence": "Прощён",
        "weekend": "Выходной",
    }
    day_rows = []
    for d in summary.get("days", []):
        ci = d.get("checked_in_at")
        co = d.get("checked_out_at")

        def _fmt_time(iso: str | None) -> str:
            if not iso:
                return ""
            try:
                parsed = dt.datetime.fromisoformat(iso)
                # attendance_payroll_summary отдаёт aware ISO — приводим
                # к локальному времени, чтобы Excel показал `10:15`, а
                # не `05:15Z`.
                if parsed.tzinfo is not None:
                    parsed = timezone.localtime(parsed)
                return parsed.strftime("%H:%M")
            except Exception:
                return iso[-8:-3] if len(iso) >= 8 else iso

        day_rows.append(
            [
                d.get("date", ""),
                weekday_ru[int(d.get("weekday", 0))],
                "Да" if d.get("is_working_day") else "Нет",
                status_ru.get(d.get("status", ""), d.get("status", "")),
                _fmt_time(ci),
                _fmt_time(co),
                int(d.get("minutes_late", 0) or 0),
                int(d.get("deduction_uzs", 0) or 0),
                d.get("note", ""),
            ]
        )
    write_sheet(
        wb,
        title="По дням",
        headers=day_headers,
        rows=day_rows,
        int_columns=(6, 7),
    )

    return workbook_response(wb, filename)


class PayrollListAttendanceApi(APIView):
    """
    Manager: `GET /api/attendance/payroll/?month=YYYY-MM` → list per-operator
    зарплатных сводок (те же поля, что и detail, только без `days[]`).
    """

    permission_classes = [IsAuthenticated, IsTeamLeadOrManager, IsAttendancePinVerified]

    def get(self, request):
        from apps.operators.models import OperatorStatus
        from .selectors import attendance_payroll_summary

        year, month = _parse_month_param(request.query_params.get("month")) or _default_month()
        operators = Operator.objects.exclude(status=OperatorStatus.INACTIVE).order_by("full_name")

        rows = []
        for op in operators:
            summary = attendance_payroll_summary(op, year, month)
            # Тонкая версия без days[] — списочный view.
            slim = {k: v for k, v in summary.items() if k != "days"}
            rows.append(slim)

        return Response(
            {
                "period": {"year": year, "month": month},
                "rows": rows,
            }
        )


class PayrollDetailAttendanceApi(APIView):
    """
    Manager: `GET /api/attendance/payroll/{operator_id}/?month=YYYY-MM
    [&format=xlsx|pdf]` — детальная сводка с days-breakdown + downloads.
    """

    permission_classes = [IsAuthenticated, IsTeamLeadOrManager, IsAttendancePinVerified]

    def get(self, request, operator_id: int):
        from .selectors import attendance_payroll_summary
        from .services import payroll_pdf_bytes, pdf_response

        operator = get_object_or_404(Operator, id=operator_id)
        year, month = _parse_month_param(request.query_params.get("month")) or _default_month()
        summary = attendance_payroll_summary(operator, year, month)

        fmt = (request.query_params.get("export") or "json").lower()
        if fmt == "xlsx":
            filename = (
                f"payroll_{operator.id}_{year}-{month:02d}.xlsx"
            )
            return _payroll_summary_to_xlsx(summary, filename)
        if fmt == "pdf":
            pdf = payroll_pdf_bytes(
                operator_name=operator.full_name,
                year=year,
                month=month,
                summary=summary,
            )
            return pdf_response(pdf, f"payroll_{operator.id}_{year}-{month:02d}.pdf")

        return Response(summary)


class MyPayrollAttendanceApi(APIView):
    """
    Operator: `GET /api/attendance/my-payroll/?month=YYYY-MM
    [&format=xlsx|pdf]` — свой зарплатный отчёт. PIN не требуется.
    """

    permission_classes = [IsAuthenticated, IsAuthenticatedAnyRole]

    def get(self, request):
        from .selectors import attendance_payroll_summary
        from .services import payroll_pdf_bytes, pdf_response

        profile = getattr(request.user, "profile", None)
        if not profile or not profile.operator_id:
            return Response(
                {"error": "Пользователь не привязан к оператору"},
                status=status.HTTP_404_NOT_FOUND,
            )
        operator = profile.operator

        year, month = _parse_month_param(request.query_params.get("month")) or _default_month()
        summary = attendance_payroll_summary(operator, year, month)

        fmt = (request.query_params.get("export") or "json").lower()
        if fmt == "xlsx":
            filename = f"my_payroll_{year}-{month:02d}.xlsx"
            return _payroll_summary_to_xlsx(summary, filename)
        if fmt == "pdf":
            pdf = payroll_pdf_bytes(
                operator_name=operator.full_name,
                year=year,
                month=month,
                summary=summary,
            )
            return pdf_response(pdf, f"my_payroll_{year}-{month:02d}.pdf")

        return Response(summary)


class DashboardSnapshotAttendanceApi(APIView):
    """
    Компактный срез посещаемости для менеджерского дашборда «Сводка дня».

    Не под PIN-гейтом (в отличие от /report/) — сюда попадают только
    сводные счётчики без имён и лог-строк:
      {"on_shift": N, "expected": M, "late_today": K}
    """

    permission_classes = [IsAuthenticated, IsTeamLeadOrManager]

    def get(self, request):
        return Response(attendance_dashboard_snapshot())
